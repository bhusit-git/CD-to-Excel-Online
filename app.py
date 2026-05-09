import streamlit as st
import pandas as pd
import re
import struct
import io
from datetime import date, datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

# --- Constants from the original script ---
THAI_MONTHS = ["", "มกราคม", "กุมภาพันธ์", "มีนาคม", "เมษายน", "พฤษภาคม", "มิถุนายน", "กรกฎาคม", "สิงหาคม", "กันยายน", "ตุลาคม", "พฤศจิกายน", "ธันวาคม"]
THAI_DIGITS = ["ศูนย์", "หนึ่ง", "สอง", "สาม", "สี่", "ห้า", "หก", "เจ็ด", "แปด", "เก้า"]
THAI_POSITIONS = ["", "สิบ", "ร้อย", "พัน", "หมื่น", "แสน"]
SMALL_PRICE = 60.0
SMALL_SUMMARY_BILLS = {"69 040870", "69 040871", "69 040872", "69 040873"}

SHEET_CONFIGS = [
    {
        "name": "ลอสันหลอดใหญ่",
        "kind": "lawson_big",
        "bill_no": "6905001",
        "customer_line": "ชื่อลูกค้า : บริษัท สห ลอว์สัน จำกัด",
        "address_line": "เลขที่ 2170 อาคารกรุงเทพทาวเวอร์ ชั้น 3 ถนนเพชรบุรีตัดใหม่ แขวงบางกะปิ เขตห้วยขวาง กรุงเทพมหานคร 10310",
        "tax_line": "เลขประจำตัวผู้เสียภาษี 105-55516-6337",
        "product_code": "04",
        "branch_ids": [
            "0829", "1089", "1360", "2880", "3089", "3204", "3208", "3255", "3534",
            "3548", "3554", "3555", "3572", "3575", "3581", "3639", "3666", "3667",
            "3674", "3675", "3676", "3709", "3728", "3733", "3735", "3747", "3750",
            "3754", "3758", "3760", "3766", "3767", "3775", "3793", "3794", "3807",
            "3814",
        ],
        "title": "รายการ : น้ำแข็งหลอดใหญ่แพ็ค 1.4 KG (ห่อ)",
    },
    {
        "name": "แฟรนไชส์(ประเวศ)",
        "kind": "franchise_big",
        "bill_no": "6905002",
        "customer_line": "ชื่อลูกค้า : บริษัท ซัสโก้มาร์เก็ตติ้ง จำกัด (สาขาที่ 00025 ประเวศ)",
        "address_line": "ที่อยู่ : 161 ถ.มอเตอร์เวย์ แขวงทับช้าง เขตสะพานสูง กรุงเทพ 10250",
        "tax_line": "เลขประจำผู้เสียภาษี : 0105539021915",
        "product_code": "04",
        "branch_ids": ["3653"],
    },
    {
        "name": "แฟรนไชส์หัวหมาก",
        "kind": "franchise_big",
        "bill_no": "6905003",
        "customer_line": "ชื่อลูกค้า : บริษัท ซัสโก้มาร์เก็ตติ้ง จำกัด (สาขาที่ 00032 หัวหมาก)",
        "address_line": "",
        "tax_line": "",
        "product_code": "04",
        "branch_ids": ["3658"],
    },
    {
        "name": "แฟรนไชส์สวนหลวง",
        "kind": "franchise_big",
        "bill_no": "6905004",
        "customer_line": "ชื่อลูกค้า : บริษัท ซัสโก้มาร์เก็ตติ้ง จำกัด (สาขาที่ 00019 สวนหลวง)",
        "address_line": "",
        "tax_line": "",
        "product_code": "04",
        "branch_ids": ["3645"],
    },
    {
        "name": "แฟรนไชสัลโก้บางพลี",
        "kind": "franchise_big",
        "bill_no": "6905005",
        "customer_line": "ชื่อลูกค้า : บริษัท ซัสโก้มาร์เก็ตติ้ง จำกัด (สาขาที 00031 บางพลี)",
        "address_line": "",
        "tax_line": "",
        "product_code": "04",
        "branch_ids": ["3663"],
    },
    {
        "name": "แฟรนไชสัสโก้บางบ่อ",
        "kind": "franchise_big",
        "bill_no": "6905006",
        "customer_line": "ชื่อลูกค้า : บริษัท ซัสโก้มาร์เก็ตติ้ง จำกัด (สาขาที 00038 บางบ่อ)",
        "address_line": "",
        "tax_line": "",
        "product_code": "04",
        "branch_ids": ["3719"],
    },
    {
        "name": "แฟรนไชสัสโก้สะพานสูง",
        "kind": "franchise_big",
        "bill_no": "6905007",
        "customer_line": "ชื่อลูกค้า : บริษัท ซัสโก้มาร์เก็ตติ้ง จำกัด (สาขาที่ 00048 ซัสโก้สะพานสูง )",
        "address_line": "",
        "tax_line": "",
        "product_code": "04",
        "branch_ids": ["3730"],
    },
    {
        "name": "แฟรนไชสัสโก้คลองสองต้นนุ่น",
        "kind": "franchise_big",
        "bill_no": "6905008",
        "customer_line": "ชื่อลูกค้า : บริษัท ซัสโก้มาร์เก็ตติ้ง จำกัด (สาขาที 00041 ซัสโก้คลองสองต้นนุ่น)",
        "address_line": "",
        "tax_line": "",
        "product_code": "04",
        "branch_ids": ["3703"],
    },
    {
        "name": "ลอสันหลอดเล็ก",
        "kind": "lawson_small_combined",
        "bill_no": "6905009",
        "customer_line": "ชื่อลูกค้า : บริษัท สห ลอว์สัน จำกัด",
        "address_line": "เลขที่ 2170 อาคารกรุงเทพทาวเวอร์ ชั้น 3 ถนนเพชรบุรีตัดใหม่ แขวงบางกะปิ เขตห้วยขวาง กรุงเทพมหานคร 10310",
        "tax_line": "เลขประจำตัวผู้เสียภาษี 105-55516-6337",
        "product_code": "06",
        "branch_ids": ["3666", "3639", "3760", "3735", "3766"],
        "title": "รายการ : น้ำแข็งหลอดเล็ก",
    },
    {
        "name": "3666 (ล.)",
        "kind": "lawson_small_single",
        "bill_no": "6905010",
        "customer_line": "ชื่อลูกค้า : บริษัท สห ลอว์สัน จำกัด",
        "address_line": "เลขที่ 2170 อาคารกรุงเทพทาวเวอร์ ชั้น 3 ถนนเพชรบุรีตัดใหม่ แขวงบางกะปิ เขตห้วยขวาง กรุงเทพมหานคร 10310",
        "tax_line": "เลขประจำตัวผู้เสียภาษี 105-55516-6337",
        "product_code": "06",
        "branch_ids": ["3666"],
        "title": "รายการ : น้ำแข็งหลอดเล็ก",
    },
    {
        "name": "3639 (ล.)",
        "kind": "lawson_small_single",
        "bill_no": "6905011",
        "customer_line": "ชื่อลูกค้า : บริษัท สห ลอว์สัน จำกัด",
        "address_line": "เลขที่ 2170 อาคารกรุงเทพทาวเวอร์ ชั้น 3 ถนนเพชรบุรีตัดใหม่ แขวงบางกะปิ เขตห้วยขวาง กรุงเทพมหานคร 10310",
        "tax_line": "เลขประจำตัวผู้เสียภาษี 105-55516-6337",
        "product_code": "06",
        "branch_ids": ["3639"],
        "title": "รายการ : น้ำแข็งหลอดเล็ก",
    },
    {
        "name": "3760 (ล.)",
        "kind": "lawson_small_single",
        "bill_no": "6905012",
        "customer_line": "ชื่อลูกค้า : บริษัท สห ลอว์สัน จำกัด",
        "address_line": "เลขที่ 2170 อาคารกรุงเทพทาวเวอร์ ชั้น 3 ถนนเพชรบุรีตัดใหม่ แขวงบางกะปิ เขตห้วยขวาง กรุงเทพมหานคร 10310",
        "tax_line": "เลขประจำตัวผู้เสียภาษี 105-55516-6337",
        "product_code": "06",
        "branch_ids": ["3760"],
        "title": "รายการ : น้ำแข็งหลอดเล็ก",
    },
    {
        "name": "3735 (ล.)",
        "kind": "lawson_small_single",
        "bill_no": "6905013",
        "customer_line": "ชื่อลูกค้า : บริษัท สห ลอว์สัน จำกัด",
        "address_line": "เลขที่ 2170 อาคารกรุงเทพทาวเวอร์ ชั้น 3 ถนนเพชรบุรีตัดใหม่ แขวงบางกะปิ เขตห้วยขวาง กรุงเทพมหานคร 10310",
        "tax_line": "เลขประจำตัวผู้เสียภาษี 105-55516-6337",
        "product_code": "06",
        "branch_ids": ["3735"],
        "title": "รายการ : น้ำแข็งหลอดเล็ก",
    },
]

# --- Helper Functions (Modified for Streamlit) ---

def parse_bill_number(raw: str) -> str:
    raw = raw.strip().replace(" ", "")
    if not raw: return ""
    try: return str(int(round(float(raw))))
    except ValueError: return raw

def numeric_key(raw: str) -> int:
    digits = re.sub(r"\D", "", str(raw))
    return int(digits) if digits else 0

def parse_ymd(raw: str) -> datetime:
    return datetime.strptime(raw.strip(), "%Y%m%d")

def format_thai_bill_date(value: date) -> str:
    return f"วันที่ {value.day} {THAI_MONTHS[value.month]} {value.year + 543}"

def month_bounds(month_text: str) -> tuple[str, str]:
    start_day = datetime.strptime(month_text + "-01", "%Y-%m-%d").date()
    next_month = (start_day.replace(day=28) + timedelta(days=4)).replace(day=1)
    end_day = next_month - timedelta(days=1)
    return start_day.strftime("%Y%m%d"), end_day.strftime("%Y%m%d")

def load_dbf_from_file(file_obj) -> list[dict]:
    # Read header
    hdr = file_obj.read(32)
    if not hdr: return []
    numrec = struct.unpack("<I", hdr[4:8])[0]
    hlen = struct.unpack("<H", hdr[8:10])[0]
    rlen = struct.unpack("<H", hdr[10:12])[0]
    
    fields = []
    while True:
        desc = file_obj.read(32)
        if not desc or desc[0] == 0x0D:
            break
        name = desc[:11].split(b"\x00", 1)[0].decode("ascii", "ignore").strip()
        typ = chr(desc[11])
        flen = desc[16]
        fields.append((name, typ, flen))
        
    offsets = []
    pos = 1
    for _, _, flen in fields:
        offsets.append(pos)
        pos += flen
        
    rows = []
    file_obj.seek(hlen)
    for _ in range(numrec):
        rec = file_obj.read(rlen)
        if not rec or rec[:1] == b"*":
            continue
        row = {}
        for (name, typ, flen), off in zip(fields, offsets):
            raw = rec[off : off + flen]
            if typ == "D":
                value = raw.decode("ascii", "ignore").strip()
            else:
                value = raw.decode("cp874", "ignore").strip(" \x00")
            row[name] = value
        rows.append(row)
    return rows

def thai_baht_text(amount: float) -> str:
    integer = int(round(amount))
    def read_under_million(num: int) -> str:
        if num == 0: return ""
        parts = []
        digits = list(map(int, str(num)))
        size = len(digits)
        for idx, digit in enumerate(digits):
            pos = size - idx - 1
            if digit == 0: continue
            if pos == 0:
                if digit == 1 and size > 1: parts.append("เอ็ด")
                else: parts.append(THAI_DIGITS[digit])
            elif pos == 1:
                if digit == 1: parts.append("สิบ")
                elif digit == 2: parts.append("ยี่สิบ")
                else: parts.append(THAI_DIGITS[digit] + "สิบ")
            else:
                parts.append(THAI_DIGITS[digit] + THAI_POSITIONS[pos])
        return "".join(parts)
    if integer == 0: return "ศูนย์บาทถ้วน"
    chunks = []
    while integer > 0:
        chunks.append(integer % 1_000_000)
        integer //= 1_000_000
    words = []
    for idx in range(len(chunks) - 1, -1, -1):
        chunk = chunks[idx]
        if chunk == 0: continue
        chunk_text = read_under_million(chunk)
        if idx > 0: words.append(chunk_text + "ล้าน")
        else: words.append(chunk_text)
    return "".join(words) + "บาทถ้วน"

def apply_base_format(ws, max_rows: int = 800) -> None:
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A12"
    col_widths = {"A": 8, "B": 18, "C": 14, "D": 12, "E": 30, "F": 12, "G": 12, "H": 14, "I": 14, "J": 14, "K": 14, "L": 14}
    for col, width in col_widths.items():
        ws.column_dimensions[col].width = width
    for row in range(1, max_rows + 1):
        for col in range(1, 13):
            cell = ws.cell(row=row, column=col)
            cell.font = Font(name="Angsana New", size=16)
            cell.alignment = Alignment(vertical="center")

def write_common_header(ws, config: dict, bill_date_str: str, bill_no: str, title: str | None) -> None:
    ws.merge_cells("A1:L1")
    ws["A1"] = "ใบวางบิล/ใบแจ้งหนี้"
    ws["A1"].font = Font(name="Angsana New", size=20, bold=True)
    ws["A1"].alignment = Alignment(horizontal="center")
    ws.merge_cells("A2:J2")
    ws["A2"] = "บริษัท ซูเปอร์ ไอซ์ จำกัด  สำนักงานใหญ่"
    ws["A2"].font = Font(name="Angsana New", size=18, bold=True)
    ws["K2"] = "ต้นฉบับ"
    ws["K2"].font = Font(name="Angsana New", size=16, bold=True)
    ws.merge_cells("A3:H3")
    ws["A3"] = "ที่อยู่ 18/39 ซอยนวมินทร์ 111 แยก 15 แขวงนวมินทร์ เขตบึงกุ่ม กรุงเทพมหานคร 10240"
    ws.merge_cells("I3:L3")
    ws["I3"] = f"เลขที่บิล {bill_no}"
    ws.merge_cells("A4:H4")
    ws["A4"] = "เลขประจำตัวผู้เสียภาษี 0105542031756"
    ws.merge_cells("I5:L5")
    ws["I5"] = bill_date_str
    ws["I5"].alignment = Alignment(horizontal="center")
    ws["I5"].font = Font(name="Angsana New", size=16, bold=True)
    ws.merge_cells("A6:L6")
    ws["A6"] = config["customer_line"]
    ws["A6"].font = Font(name="Angsana New", size=16, bold=True)
    ws.merge_cells("A7:L7")
    ws["A7"] = config["address_line"]
    ws.merge_cells("A8:L8")
    ws["A8"] = config["tax_line"]
    if title:
        ws.merge_cells("A9:L9")
        ws["A9"] = title
        ws["A9"].font = Font(name="Angsana New", size=16, bold=True)

def clean_branch_name(raw: str) -> str:
    raw = raw.strip()
    raw = re.sub(r"^P\d+\s*สาขา\s*", "", raw)
    return raw

def build_trans_rows(customers: dict, trans_rows: list, branch_ids: list | None, product_code: str, p_start: str, p_end: str) -> list[dict]:
    branch_set = set(branch_ids or [])
    rows = []
    for row in trans_rows:
        cust_id = row.get("CUST_VEND", "")
        if branch_set and cust_id not in branch_set: continue
        if row.get("PROD_CODE") != product_code: continue
        date_raw = row.get("DATE", "")
        if not (p_start <= date_raw <= p_end): continue
        cust = customers.get(cust_id, {})
        total = float(row.get("AMT") or 0)
        amount = round(total / 1.07, 12)
        vat = round(total - amount, 12)
        branch_name = clean_branch_name(cust.get("SH_NAME") or cust.get("SHIP_NAME") or cust.get("NAME") or "")
        rows.append({
            "date": parse_ymd(date_raw),
            "branch": f"P{cust_id}",
            "bill_no": parse_bill_number(row.get("BILL_NO", "")),
            "branch_name": branch_name,
            "qty": float(row.get("QTY") or 0),
            "price": float(row.get("PRICE") or 0),
            "amount": amount,
            "vat": vat,
            "total": total,
            "product_name": "หลอดใหญ่" if product_code == "04" else branch_name,
        })
    rows.sort(key=lambda r: (r["branch_name"], r["date"], r["branch"], r["bill_no"]))
    for idx, row in enumerate(rows, start=1): row["seq"] = idx
    return rows

def build_small_bill_rows(customers: dict, bill_rows: list, branch_ids: list | None, p_start: str, p_end: str) -> list[dict]:
    branch_set = set(branch_ids or [])
    rows = []
    for row in bill_rows:
        cust_id = row.get("CUST_VEND", "")
        if branch_set and cust_id not in branch_set: continue
        date_raw = row.get("REF_DATE", "")
        if not (p_start <= date_raw <= p_end): continue
        if row.get("NO", "") in SMALL_SUMMARY_BILLS: continue
        amount = float(row.get("BAL_AMT") or 0)
        vat = float(row.get("VAT_AMT") or 0)
        total = round(amount + vat, 2)
        if total <= 0: continue
        cust = customers.get(cust_id, {})
        branch_name = clean_branch_name(cust.get("SH_NAME") or cust.get("SHIP_NAME") or cust.get("NAME") or "")
        qty = int(round(total / SMALL_PRICE))
        rows.append({
            "date": parse_ymd(date_raw),
            "branch": f"P{cust_id}",
            "bill_no": parse_bill_number(row.get("NO", "")),
            "branch_name": branch_name,
            "qty": qty,
            "price": SMALL_PRICE,
            "amount": amount,
            "vat": vat,
            "total": total,
            "product_name": branch_name,
        })
    rows.sort(key=lambda r: (numeric_key(r["branch"]), r["date"], numeric_key(r["bill_no"])))
    for idx, row in enumerate(rows, start=1): row["seq"] = idx
    return rows

def make_border():
    thin = Side(style="thin", color="000000")
    return Border(left=thin, right=thin, top=thin, bottom=thin)

def write_lawson_big_sheet(ws, config: dict, rows: list[dict], bill_date_str: str) -> None:
    apply_base_format(ws, max_rows=900)
    write_common_header(ws, config, bill_date_str, config["bill_no"], config["title"])
    headers = ["ลำดับ", "วันที่", "เลขที่สาขา", "เลขที่บิล", "สาขา", "จำนวน", "ราคา/หน่วย", "จำนวนเงิน", "ภาษีมูลค่าเพิ่ม", "เงินรวมทั้งสิ้น"]
    border = make_border()
    fill = PatternFill("solid", fgColor="D9EAD3")
    row_idx = 11
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row_idx, column=col_idx, value=header)
        cell.font = Font(name="Angsana New", size=16, bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
        cell.fill = fill
    data_start = 12
    for row_idx, record in enumerate(rows, start=data_start):
        values = [record["seq"], record["date"], record["branch"], record["bill_no"], record["branch_name"], record["qty"], record["price"], f"=ROUND(F{row_idx}*G{row_idx}/1.07,2)", f"=ROUND(J{row_idx}-H{row_idx},2)", f"=ROUND(F{row_idx}*G{row_idx},2)"]
        for col_idx, value in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = border
            if col_idx == 2:
                cell.number_format = "dd/mm/yyyy"
                cell.alignment = Alignment(horizontal="center")
            elif col_idx in (1, 6):
                cell.number_format = "0"
                cell.alignment = Alignment(horizontal="center")
            elif col_idx == 7:
                cell.number_format = "0.00"
                cell.alignment = Alignment(horizontal="center")
            elif col_idx in (8, 9, 10):
                cell.number_format = "#,##0.00"
                cell.alignment = Alignment(horizontal="right")
            elif col_idx in (3, 4): cell.alignment = Alignment(horizontal="center")
    total_row = data_start + len(rows)
    total_grand = sum(r["total"] for r in rows)
    ws.merge_cells(start_row=total_row, start_column=1, end_row=total_row, end_column=2)
    ws[f"A{total_row}"] = "รวมทั้งหมด"
    ws[f"C{total_row}"] = "IN64002102"
    ws.merge_cells(start_row=total_row, start_column=4, end_row=total_row, end_column=7)
    ws[f"D{total_row}"] = thai_baht_text(total_grand)
    ws[f"H{total_row}"] = f"=SUM(H{data_start}:H{total_row - 1})"
    ws[f"I{total_row}"] = f"=SUM(I{data_start}:I{total_row - 1})"
    ws[f"J{total_row}"] = f"=SUM(J{data_start}:J{total_row - 1})"
    for col_idx in range(1, 11):
        cell = ws.cell(row=total_row, column=col_idx)
        cell.border = border
        if col_idx in (8, 9, 10):
            cell.number_format = "#,##0.00"
            cell.font = Font(name="Angsana New", size=16, bold=True)
    footer_row = total_row + 3
    ws.merge_cells(start_row=footer_row, start_column=1, end_row=footer_row, end_column=4)
    ws.cell(row=footer_row, column=1, value="ผู้วางบิล ……………………………………..")
    ws.merge_cells(start_row=footer_row, start_column=7, end_row=footer_row, end_column=9)
    ws.cell(row=footer_row, column=7, value="ผู้รับวางบิล  .................................................")
    ws.merge_cells(start_row=footer_row + 1, start_column=7, end_row=footer_row + 1, end_column=9)
    ws.cell(row=footer_row + 1, column=7, value="วันที่รับวางบิล………………………..")
    ws.cell(row=footer_row, column=10, value="วันที่วางบิล.................................")
    ws.cell(row=footer_row + 1, column=10, value="วันที่จ่ายชำระ.............................")

def write_narrow_sheet(ws, config: dict, rows: list[dict], combined_small: bool, bill_date_str: str) -> None:
    apply_base_format(ws, max_rows=400)
    write_common_header(ws, config, bill_date_str, config["bill_no"], config.get("title"))
    border = make_border()
    fill = PatternFill("solid", fgColor="D9EAD3")
    headers = ["ลำดับ", "วันที่", "เลขที่บิล", "รหัสสาขา", "ประเภท", "จำนวน", "ราคา", "จำนวนเงิน", "ภาษีมูลค่าเพิ่ม", "เงินรวมทั้งสิ้น"]
    row_idx = 10
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row_idx, column=col_idx, value=header)
        cell.font = Font(name="Angsana New", size=16, bold=True)
        cell.alignment = Alignment(horizontal="center")
        cell.border = border
        cell.fill = fill
    data_start = 11
    for row_idx, record in enumerate(rows, start=data_start):
        label = record["branch_name"] if combined_small or config["kind"] == "lawson_small_single" else "หลอดใหญ่"
        values = [record["seq"], record["date"], record["bill_no"], record["branch"], label, record["qty"], record["price"], f"=ROUND(F{row_idx}*G{row_idx}/1.07,2)", f"=ROUND(J{row_idx}-H{row_idx},2)", f"=ROUND(F{row_idx}*G{row_idx},2)"]
        for col_idx, value in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = border
            if col_idx == 2:
                cell.number_format = "dd/mm/yyyy"
                cell.alignment = Alignment(horizontal="center")
            elif col_idx in (1, 6):
                cell.number_format = "0"
                cell.alignment = Alignment(horizontal="center")
            elif col_idx == 7:
                cell.number_format = "0.00"
                cell.alignment = Alignment(horizontal="center")
            elif col_idx in (8, 9, 10):
                cell.number_format = "#,##0.00"
                cell.alignment = Alignment(horizontal="right")
            elif col_idx in (3, 4): cell.alignment = Alignment(horizontal="center")
    total_row = data_start + len(rows)
    total_grand = sum(r["total"] for r in rows)
    ws[f"B{total_row}"] = "รวม"
    ws[f"C{total_row}"] = thai_baht_text(total_grand)
    ws[f"F{total_row}"] = f"=SUM(F{data_start}:F{total_row - 1})"
    ws[f"H{total_row}"] = f"=SUM(H{data_start}:H{total_row - 1})"
    ws[f"I{total_row}"] = f"=SUM(I{data_start}:I{total_row - 1})"
    ws[f"J{total_row}"] = f"=SUM(J{data_start}:J{total_row - 1})"
    for col_idx in range(1, 11):
        cell = ws.cell(row=total_row, column=col_idx)
        cell.border = border
        if col_idx in (8, 9, 10): cell.number_format = "#,##0.00"
    footer_row = total_row + 4
    ws.cell(row=footer_row, column=1, value="ลงชื่อ")
    ws.cell(row=footer_row, column=2, value="............................................... ผู้วางบิล")
    ws.cell(row=footer_row, column=7, value="ลงชื่อ  ............................................................")
    ws.cell(row=footer_row, column=10, value="ผู้รับวางบิล")
    ws.cell(row=footer_row + 2, column=1, value="วันที่วางบิล.................................")
    ws.cell(row=footer_row + 2, column=7, value="วันที่รับวางบิล .................................................")
    ws.cell(row=footer_row + 4, column=1, value="วันที่จ่ายชำระ.............................")

# --- Streamlit UI ---

st.set_page_config(page_title="Super Ice Billing System", layout="wide")
st.title("❄️ ระบบสร้างใบวางบิล Super Ice (Lawson/Franchise)")

col1, col2 = st.columns(2)
with col1:
    month_val = st.text_input("เดือนข้อมูล (YYYY-MM)", value=datetime.now().strftime("%Y-%m"))
    bill_date = st.date_input("วันที่วางบิล", value=date.today())

with col2:
    st.info("อัปโหลดไฟล์ .DBF จากโปรแกรมบัญชี")
    big_files = st.file_uploader("ข้อมูลหลอดใหญ่/แฟรนไชส์ (MCUST.DBF, ATRANS.DBF, ABILLNO.DBF)", accept_multiple_files=True)
    small_files = st.file_uploader("ข้อมูลหลอดเล็ก (Optional - LAWSON)", accept_multiple_files=True)

if st.button("🚀 สร้างไฟล์วางบิล", type="primary"):
    if not big_files:
        st.error("กรุณาอัปโหลดไฟล์ข้อมูลหลอดใหญ่")
    else:
        try:
            # Load Data
            def get_dbf_data(files):
                data = {}
                for f in files:
                    name = f.name.upper()
                    if name == "MCUST.DBF": 
                        content = f.read()
                        data["customers"] = load_dbf_from_file(io.BytesIO(content))
                    elif name == "ATRANS.DBF": 
                        content = f.read()
                        data["trans"] = load_dbf_from_file(io.BytesIO(content))
                    elif name == "ABILLNO.DBF": 
                        content = f.read()
                        data["bill"] = load_dbf_from_file(io.BytesIO(content))
                return data

            big_data = get_dbf_data(big_files)
            small_data = get_dbf_data(small_files) if small_files else big_data
            
            if "customers" not in big_data or "trans" not in big_data:
                st.error("ไฟล์ข้อมูลหลอดใหญ่ไม่ครบ (ต้องมี MCUST.DBF และ ATRANS.DBF)")
            else:
                p_start, p_end = month_bounds(month_val)
                bill_date_str = format_thai_bill_date(bill_date)
                
                big_cust_dict = {row["ID"]: row for row in big_data["customers"]}
                
                wb = Workbook()
                wb.remove(wb.active)
                
                for config in SHEET_CONFIGS:
                    ws = wb.create_sheet(config["name"])
                    if config["kind"].startswith("lawson_small"):
                        # Use small data if available, fallback to big
                        target_bill = small_data.get("bill", big_data.get("bill", []))
                        target_cust = {row["ID"]: row for row in small_data.get("customers", big_data["customers"])}
                        rows = build_small_bill_rows(target_cust, target_bill, config["branch_ids"], p_start, p_end)
                    else:
                        rows = build_trans_rows(big_cust_dict, big_data["trans"], config["branch_ids"], config["product_code"], p_start, p_end)
                    
                    if config["kind"] == "lawson_big": write_lawson_big_sheet(ws, config, rows, bill_date_str)
                    else: write_narrow_sheet(ws, config, rows, config["kind"].endswith("combined"), bill_date_str)
                
                # Save to memory
                output = io.BytesIO()
                wb.save(output)
                output.seek(0)
                
                st.success("✅ สร้างไฟล์สำเร็จ!")
                st.download_button(
                    label="📥 ดาวน์โหลดไฟล์ Excel",
                    data=output,
                    file_name=f"วางบิล_{month_val}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
        except Exception as e:
            st.error(f"เกิดข้อผิดพลาด: {str(e)}")
