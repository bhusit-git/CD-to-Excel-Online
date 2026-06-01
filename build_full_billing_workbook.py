from __future__ import annotations

import argparse
import re
import struct
from datetime import date
from datetime import datetime
from datetime import timedelta
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


SCRIPT_DIR = Path(__file__).resolve().parent
DEFAULT_BASE_DIR = SCRIPT_DIR
DEFAULT_OUTPUT_XLSX = SCRIPT_DIR / "outputs" / "วางบิล_อัปเดตจาก_SI-2568.xlsx"
BASE_DIR = DEFAULT_BASE_DIR
OUTPUT_XLSX = DEFAULT_OUTPUT_XLSX
BILL_DATE = "วันที่ 11 พฤษภาคม 2569"
SMALL_PRICE = 60.0
SMALL_SUMMARY_BILLS = {"69 040870", "69 040871", "69 040872", "69 040873"}
PERIOD_START = "20260401"
PERIOD_END = "20260430"
BODY_FONT_SIZE = 18
SUBTITLE_FONT_SIZE = 20
TITLE_FONT_SIZE = 22
THAI_MONTHS = [
    "",
    "มกราคม",
    "กุมภาพันธ์",
    "มีนาคม",
    "เมษายน",
    "พฤษภาคม",
    "มิถุนายน",
    "กรกฎาคม",
    "สิงหาคม",
    "กันยายน",
    "ตุลาคม",
    "พฤศจิกายน",
    "ธันวาคม",
]


SHEET_CONFIGS = [
    {
        "name": "ลอสันหลอดใหญ่",
        "kind": "lawson_big",
        "bill_no": "6905001",
        "customer_line": "ชื่อลูกค้า : บริษัท สห ลอว์สัน จำกัด",
        "address_line": "เลขที่ 2170 อาคารกรุงเทพทาวเวอร์ ชั้น 3 ถนนเพชรบุรีตัดใหม่ แขวงบางกะปิ เขตห้วยขวาง กรุงเทพมหานคร 10310",
        "tax_line": "เลขประจำตัวผู้เสียภาษี 105-55516-6337",
        "product_code": "04",
        "branch_ids": [
            "0829", "1089", "1360", "2880", "3089", "3204", "3208", "3255", "3534",
            "3548", "3554", "3555", "3572", "3575", "3581", "3639", "3666", "3667",
            "3674", "3675", "3676", "3709", "3728", "3733", "3735", "3747", "3750",
            "3754", "3758", "3760", "3766", "3767", "3775", "3793", "3794", "3807",
            "3814",
        ],
        "title": "รายการ : น้ำแข็งหลอดใหญ่แพ็ค 1.4 KG (ห่อ)",
    },
    {
        "name": "แฟรนไชส์(ประเวศ)",
        "kind": "franchise_big",
        "bill_no": "6905002",
        "customer_line": "ชื่อลูกค้า : บริษัท ซัสโก้มาร์เก็ตติ้ง จำกัด (สาขาที่ 00025 ประเวศ)",
        "address_line": "ที่อยู่ : 161 ถ.มอเตอร์เวย์ แขวงทับช้าง เขตสะพานสูง กรุงเทพ 10250",
        "tax_line": "เลขประจำผู้เสียภาษี : 0105539021915",
        "product_code": "04",
        "branch_ids": ["3653"],
    },
    {
        "name": "แฟรนไชส์หัวหมาก",
        "kind": "franchise_big",
        "bill_no": "6905003",
        "customer_line": "ชื่อลูกค้า : บริษัท ซัสโก้มาร์เก็ตติ้ง จำกัด (สาขาที่ 00032 หัวหมาก)",
        "address_line": "",
        "tax_line": "",
        "product_code": "04",
        "branch_ids": ["3658"],
    },
    {
        "name": "แฟรนไชส์สวนหลวง",
        "kind": "franchise_big",
        "bill_no": "6905004",
        "customer_line": "ชื่อลูกค้า : บริษัท ซัสโก้มาร์เก็ตติ้ง จำกัด (สาขาที่ 00019 สวนหลวง)",
        "address_line": "",
        "tax_line": "",
        "product_code": "04",
        "branch_ids": ["3645"],
    },
    {
        "name": "แฟรนไชสัลโก้บางพลี",
        "kind": "franchise_big",
        "bill_no": "6905005",
        "customer_line": "ชื่อลูกค้า : บริษัท ซัสโก้มาร์เก็ตติ้ง จำกัด (สาขาที 00031 บางพลี)",
        "address_line": "",
        "tax_line": "",
        "product_code": "04",
        "branch_ids": ["3663"],
    },
    {
        "name": "แฟรนไชสัสโก้บางบ่อ",
        "kind": "franchise_big",
        "bill_no": "6905006",
        "customer_line": "ชื่อลูกค้า : บริษัท ซัสโก้มาร์เก็ตติ้ง จำกัด (สาขาที 00038 บางบ่อ)",
        "address_line": "",
        "tax_line": "",
        "product_code": "04",
        "branch_ids": ["3719"],
    },
    {
        "name": "แฟรนไชสัสโก้สะพานสูง",
        "kind": "franchise_big",
        "bill_no": "6905007",
        "customer_line": "ชื่อลูกค้า : บริษัท ซัสโก้มาร์เก็ตติ้ง จำกัด (สาขาที่ 00048 ซัสโก้สะพานสูง )",
        "address_line": "",
        "tax_line": "",
        "product_code": "04",
        "branch_ids": ["3730"],
    },
    {
        "name": "แฟรนไชสัสโก้คลองสองต้นนุ่น",
        "kind": "franchise_big",
        "bill_no": "6905008",
        "customer_line": "ชื่อลูกค้า : บริษัท ซัสโก้มาร์เก็ตติ้ง จำกัด (สาขาที 00041 ซัสโก้คลองสองต้นนุ่น)",
        "address_line": "",
        "tax_line": "",
        "product_code": "04",
        "branch_ids": ["3703"],
    },
    {
        "name": "ลอสันหลอดเล็ก",
        "kind": "lawson_small_combined",
        "bill_no": "6905009",
        "customer_line": "ชื่อลูกค้า : บริษัท สห ลอว์สัน จำกัด",
        "address_line": "เลขที่ 2170 อาคารกรุงเทพทาวเวอร์ ชั้น 3 ถนนเพชรบุรีตัดใหม่ แขวงบางกะปิ เขตห้วยขวาง กรุงเทพมหานคร 10310",
        "tax_line": "เลขประจำตัวผู้เสียภาษี 105-55516-6337",
        "product_code": "06",
        "branch_ids": ["3666", "3639", "3760", "3735", "3766"],
        "title": "รายการ : น้ำแข็งหลอดเล็ก",
    },
    {
        "name": "3666 (ล.)",
        "kind": "lawson_small_single",
        "bill_no": "6905010",
        "customer_line": "ชื่อลูกค้า : บริษัท สห ลอว์สัน จำกัด",
        "address_line": "เลขที่ 2170 อาคารกรุงเทพทาวเวอร์ ชั้น 3 ถนนเพชรบุรีตัดใหม่ แขวงบางกะปิ เขตห้วยขวาง กรุงเทพมหานคร 10310",
        "tax_line": "เลขประจำตัวผู้เสียภาษี 105-55516-6337",
        "product_code": "06",
        "branch_ids": ["3666"],
        "title": "รายการ : น้ำแข็งหลอดเล็ก",
    },
    {
        "name": "3639 (ล.)",
        "kind": "lawson_small_single",
        "bill_no": "6905011",
        "customer_line": "ชื่อลูกค้า : บริษัท สห ลอว์สัน จำกัด",
        "address_line": "เลขที่ 2170 อาคารกรุงเทพทาวเวอร์ ชั้น 3 ถนนเพชรบุรีตัดใหม่ แขวงบางกะปิ เขตห้วยขวาง กรุงเทพมหานคร 10310",
        "tax_line": "เลขประจำตัวผู้เสียภาษี 105-55516-6337",
        "product_code": "06",
        "branch_ids": ["3639"],
        "title": "รายการ : น้ำแข็งหลอดเล็ก",
    },
    {
        "name": "3760 (ล.)",
        "kind": "lawson_small_single",
        "bill_no": "6905012",
        "customer_line": "ชื่อลูกค้า : บริษัท สห ลอว์สัน จำกัด",
        "address_line": "เลขที่ 2170 อาคารกรุงเทพทาวเวอร์ ชั้น 3 ถนนเพชรบุรีตัดใหม่ แขวงบางกะปิ เขตห้วยขวาง กรุงเทพมหานคร 10310",
        "tax_line": "เลขประจำตัวผู้เสียภาษี 105-55516-6337",
        "product_code": "06",
        "branch_ids": ["3760"],
        "title": "รายการ : น้ำแข็งหลอดเล็ก",
    },
    {
        "name": "3735 (ล.)",
        "kind": "lawson_small_single",
        "bill_no": "6905013",
        "customer_line": "ชื่อลูกค้า : บริษัท สห ลอว์สัน จำกัด",
        "address_line": "เลขที่ 2170 อาคารกรุงเทพทาวเวอร์ ชั้น 3 ถนนเพชรบุรีตัดใหม่ แขวงบางกะปิ เขตห้วยขวาง กรุงเทพมหานคร 10310",
        "tax_line": "เลขประจำตัวผู้เสียภาษี 105-55516-6337",
        "product_code": "06",
        "branch_ids": ["3735"],
        "title": "รายการ : น้ำแข็งหลอดเล็ก",
    },
]


THAI_DIGITS = ["ศูนย์", "หนึ่ง", "สอง", "สาม", "สี่", "ห้า", "หก", "เจ็ด", "แปด", "เก้า"]
THAI_POSITIONS = ["", "สิบ", "ร้อย", "พัน", "หมื่น", "แสน"]


def parse_bill_number(raw: str) -> str:
    raw = raw.strip().replace(" ", "")
    if not raw:
        return ""
    try:
        return str(int(round(float(raw))))
    except ValueError:
        return raw


def numeric_key(raw: str) -> int:
    digits = re.sub(r"\D", "", str(raw))
    return int(digits) if digits else 0


def parse_ymd(raw: str) -> datetime:
    return datetime.strptime(raw.strip(), "%Y%m%d")


def format_thai_bill_date(value: date) -> str:
    return f"วันที่ {value.day} {THAI_MONTHS[value.month]} {value.year + 543}"


def month_bounds(month_text: str) -> tuple[str, str]:
    start_day = datetime.strptime(month_text + "-01", "%Y-%m-%d").date()
    next_month = (start_day.replace(day=28) + timedelta(days=4)).replace(day=1)
    end_day = next_month - timedelta(days=1)
    return start_day.strftime("%Y%m%d"), end_day.strftime("%Y%m%d")


def resolve_source_dir(input_path: Path) -> Path:
    input_path = input_path.expanduser().resolve()
    if (input_path / "MCUST.DBF").exists() and (input_path / "ATRANS.DBF").exists():
        return input_path

    candidates = []
    for child in input_path.iterdir():
        if not child.is_dir():
            continue
        if (child / "MCUST.DBF").exists() and (child / "ATRANS.DBF").exists():
            candidates.append(child)
    if not candidates:
        raise FileNotFoundError(f"ไม่พบโฟลเดอร์ข้อมูลที่มีไฟล์ DBF ใน {input_path}")
    preferred_names = {"LAWSON", "SI-2568"}
    candidates.sort(
        key=lambda p: (p.name.upper() not in preferred_names, -p.stat().st_mtime, p.name.upper())
    )
    return candidates[0]


def resolve_preferred_source_dir(root_input: Path, preferred_name: str) -> Path:
    root_input = root_input.expanduser().resolve()
    preferred_upper = preferred_name.upper()

    if preferred_upper in {"SI", "SI-*", "SI-XXXX"}:
        if root_input.is_dir() and root_input.name.upper().startswith("SI-") and (root_input / "MCUST.DBF").exists():
            return root_input

        si_candidates = []
        for child in root_input.iterdir():
            if child.is_dir() and child.name.upper().startswith("SI-") and (child / "MCUST.DBF").exists():
                si_candidates.append(child)
        if si_candidates:
            si_candidates.sort(key=lambda p: (-p.stat().st_mtime, p.name.upper()))
            return si_candidates[0]

    direct_child = root_input / preferred_name
    if direct_child.is_dir() and (direct_child / "MCUST.DBF").exists():
        return direct_child

    if root_input.name.upper() == preferred_upper and (root_input / "MCUST.DBF").exists():
        return root_input

    candidates = []
    for child in root_input.iterdir():
        if child.is_dir() and child.name.upper() == preferred_upper and (child / "MCUST.DBF").exists():
            candidates.append(child)
    if candidates:
        candidates.sort(key=lambda p: (-p.stat().st_mtime, p.name.upper()))
        return candidates[0]

    return resolve_source_dir(root_input)


def load_dbf(path: Path) -> list[dict]:
    with path.open("rb") as f:
        hdr = f.read(32)
        numrec = struct.unpack("<I", hdr[4:8])[0]
        hlen = struct.unpack("<H", hdr[8:10])[0]
        rlen = struct.unpack("<H", hdr[10:12])[0]
        fields = []
        while True:
            desc = f.read(32)
            if not desc or desc[0] == 0x0D:
                break
            name = desc[:11].split(b"\x00", 1)[0].decode("ascii", "ignore").strip()
            typ = chr(desc[11])
            flen = desc[16]
            fields.append((name, typ, flen))
        offsets = []
        pos = 1
        for _, _, flen in fields:
            offsets.append(pos)
            pos += flen
        rows = []
        f.seek(hlen)
        for _ in range(numrec):
            rec = f.read(rlen)
            if not rec or rec[:1] == b"*":
                continue
            row = {}
            for (name, typ, flen), off in zip(fields, offsets):
                raw = rec[off : off + flen]
                if typ == "D":
                    value = raw.decode("ascii", "ignore").strip()
                else:
                    value = raw.decode("cp874", "ignore").strip(" \x00")
                row[name] = value
            rows.append(row)
    return rows


def thai_baht_text(amount: float) -> str:
    total_satang = int(round((amount or 0) * 100))
    baht = total_satang // 100
    satang = total_satang % 100

    def read_under_million(num: int) -> str:
        if num == 0:
            return ""
        parts = []
        digits = list(map(int, str(num)))
        size = len(digits)
        for idx, digit in enumerate(digits):
            pos = size - idx - 1
            if digit == 0:
                continue
            if pos == 0:
                if digit == 1 and size > 1:
                    parts.append("เอ็ด")
                else:
                    parts.append(THAI_DIGITS[digit])
            elif pos == 1:
                if digit == 1:
                    parts.append("สิบ")
                elif digit == 2:
                    parts.append("ยี่สิบ")
                else:
                    parts.append(THAI_DIGITS[digit] + "สิบ")
            else:
                parts.append(THAI_DIGITS[digit] + THAI_POSITIONS[pos])
        return "".join(parts)

    def read_number(num: int) -> str:
        if num == 0:
            return "ศูนย์"

        chunks = []
        while num > 0:
            chunks.append(num % 1_000_000)
            num //= 1_000_000

        words = []
        for idx in range(len(chunks) - 1, -1, -1):
            chunk = chunks[idx]
            if chunk == 0:
                continue
            chunk_text = read_under_million(chunk)
            if idx > 0:
                words.append(chunk_text + "ล้าน")
            else:
                words.append(chunk_text)
        return "".join(words)

    baht_text = f"{read_number(baht)}บาท"
    if satang == 0:
        return f"{baht_text}ถ้วน"
    return f"{baht_text}{read_number(satang)}สตางค์"


def apply_base_format(ws, max_rows: int = 800) -> None:
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A12"
    col_widths = {
        "A": 8,
        "B": 18,
        "C": 14,
        "D": 12,
        "E": 30,
        "F": 12,
        "G": 12,
        "H": 14,
        "I": 14,
        "J": 14,
        "K": 14,
        "L": 14,
    }
    for col, width in col_widths.items():
        ws.column_dimensions[col].width = width
    for row in range(1, max_rows + 1):
        for col in range(1, 13):
            cell = ws.cell(row=row, column=col)
            cell.font = Font(name="Angsana New", size=BODY_FONT_SIZE)
            cell.alignment = Alignment(vertical="center")


def write_common_header(ws, customer_line: str, address_line: str, tax_line: str, bill_no: str, title: str | None) -> None:
    ws.merge_cells("A1:L1")
    ws["A1"] = "ใบวางบิล/ใบแจ้งหนี้"
    ws["A1"].font = Font(name="Angsana New", size=TITLE_FONT_SIZE, bold=True)
    ws["A1"].alignment = Alignment(horizontal="center")

    ws.merge_cells("A2:J2")
    ws["A2"] = "บริษัท ซูเปอร์ ไอซ์ จำกัด  สำนักงานใหญ่"
    ws["A2"].font = Font(name="Angsana New", size=SUBTITLE_FONT_SIZE, bold=True)
    ws["K2"] = "ต้นฉบับ"
    ws["K2"].font = Font(name="Angsana New", size=BODY_FONT_SIZE, bold=True)

    ws.merge_cells("A3:H3")
    ws["A3"] = "ที่อยู่ 18/39 ซอยนวมินทร์ 111 แยก 15 แขวงนวมินทร์ เขตบึงกุ่ม กรุงเทพมหานคร 10240"
    ws.merge_cells("I3:L3")
    ws["I3"] = f"เลขที่บิล {bill_no}"
    ws["I3"].alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells("A4:H4")
    ws["A4"] = "เลขประจำตัวผู้เสียภาษี 0105542031756"

    ws.merge_cells("I5:L5")
    ws["I5"] = BILL_DATE
    ws["I5"].alignment = Alignment(horizontal="center")
    ws["I5"].font = Font(name="Angsana New", size=BODY_FONT_SIZE, bold=True)

    ws.merge_cells("A6:L6")
    ws["A6"] = customer_line
    ws["A6"].font = Font(name="Angsana New", size=BODY_FONT_SIZE, bold=True)

    ws.merge_cells("A7:L7")
    ws["A7"] = address_line

    ws.merge_cells("A8:L8")
    ws["A8"] = tax_line

    if title:
        ws.merge_cells("A9:L9")
        ws["A9"] = title
        ws["A9"].font = Font(name="Angsana New", size=BODY_FONT_SIZE, bold=True)


def load_context(source_dir: Path) -> tuple[dict[str, dict], list[dict], list[dict]]:
    customers = {row["ID"]: row for row in load_dbf(source_dir / "MCUST.DBF")}
    trans_rows = load_dbf(source_dir / "ATRANS.DBF")
    bill_rows = load_dbf(source_dir / "ABILLNO.DBF")
    return customers, trans_rows, bill_rows


def clean_branch_name(raw: str) -> str:
    raw = raw.strip()
    raw = re.sub(r"^P\d+\s*สาขา\s*", "", raw)
    return raw


def build_trans_rows(customers: dict[str, dict], trans_rows: list[dict], branch_ids: list[str] | None, product_code: str) -> list[dict]:
    branch_set = set(branch_ids or [])
    rows = []
    for row in trans_rows:
        cust_id = row.get("CUST_VEND", "")
        if branch_set and cust_id not in branch_set:
            continue
        if row.get("PROD_CODE") != product_code:
            continue
        date_raw = row.get("DATE", "")
        if not (PERIOD_START <= date_raw <= PERIOD_END):
            continue

        cust = customers.get(cust_id, {})
        total = float(row.get("AMT") or 0)
        amount = round(total / 1.07, 12)
        vat = round(total - amount, 12)
        branch_name = clean_branch_name(cust.get("SH_NAME") or cust.get("SHIP_NAME") or cust.get("NAME") or "")
        rows.append(
            {
                "date": parse_ymd(date_raw),
                "branch": f"P{cust_id}",
                "bill_no": parse_bill_number(row.get("BILL_NO", "")),
                "branch_name": branch_name,
                "qty": float(row.get("QTY") or 0),
                "price": float(row.get("PRICE") or 0),
                "amount": amount,
                "vat": vat,
                "total": total,
                "product_name": "หลอดใหญ่" if product_code == "04" else branch_name,
            }
        )
    rows.sort(key=lambda r: (r["branch_name"], r["date"], r["branch"], r["bill_no"]))
    for idx, row in enumerate(rows, start=1):
        row["seq"] = idx
    return rows


def build_small_bill_rows(customers: dict[str, dict], bill_rows: list[dict], branch_ids: list[str] | None) -> list[dict]:
    branch_set = set(branch_ids or [])
    rows = []
    for row in bill_rows:
        cust_id = row.get("CUST_VEND", "")
        if branch_set and cust_id not in branch_set:
            continue
        date_raw = row.get("REF_DATE", "")
        if not (PERIOD_START <= date_raw <= PERIOD_END):
            continue
        if row.get("NO", "") in SMALL_SUMMARY_BILLS:
            continue

        amount = float(row.get("BAL_AMT") or 0)
        vat = float(row.get("VAT_AMT") or 0)
        total = round(amount + vat, 2)
        if total <= 0:
            continue

        cust = customers.get(cust_id, {})
        branch_name = clean_branch_name(cust.get("SH_NAME") or cust.get("SHIP_NAME") or cust.get("NAME") or "")
        qty = int(round(total / SMALL_PRICE))
        rows.append(
            {
                "date": parse_ymd(date_raw),
                "branch": f"P{cust_id}",
                "bill_no": parse_bill_number(row.get("NO", "")),
                "branch_name": branch_name,
                "qty": qty,
                "price": SMALL_PRICE,
                "amount": amount,
                "vat": vat,
                "total": total,
                "product_name": branch_name,
            }
        )
    rows.sort(key=lambda r: (numeric_key(r["branch"]), r["date"], numeric_key(r["bill_no"])))
    for idx, row in enumerate(rows, start=1):
        row["seq"] = idx
    return rows


def make_border():
    thin = Side(style="thin", color="000000")
    return Border(left=thin, right=thin, top=thin, bottom=thin)


def write_lawson_big_sheet(ws, config: dict, rows: list[dict]) -> None:
    apply_base_format(ws, max_rows=900)
    write_common_header(ws, config["customer_line"], config["address_line"], config["tax_line"], config["bill_no"], config["title"])
    headers = ["ลำดับ", "วันที่", "เลขที่สาขา", "เลขที่บิล", "สาขา", "จำนวน", "ราคา/หน่วย", "จำนวนเงิน", "ภาษีมูลค่าเพิ่ม", "เงินรวมทั้งสิ้น"]
    border = make_border()
    fill = PatternFill("solid", fgColor="D9EAD3")
    row_idx = 11
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row_idx, column=col_idx, value=header)
        cell.font = Font(name="Angsana New", size=BODY_FONT_SIZE, bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
        cell.fill = fill
    data_start = 12
    for row_idx, record in enumerate(rows, start=data_start):
        values = [
            record["seq"],
            record["date"],
            record["branch"],
            record["bill_no"],
            record["branch_name"],
            record["qty"],
            record["price"],
            f"=ROUND(F{row_idx}*G{row_idx}/1.07,2)",
            f"=ROUND(J{row_idx}-H{row_idx},2)",
            f"=ROUND(F{row_idx}*G{row_idx},2)",
        ]
        for col_idx, value in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = border
            if col_idx == 2:
                cell.number_format = "dd/mm/yyyy"
                cell.alignment = Alignment(horizontal="center")
            elif col_idx in (1, 6):
                cell.number_format = "0"
                cell.alignment = Alignment(horizontal="center")
            elif col_idx == 7:
                cell.number_format = "0.00"
                cell.alignment = Alignment(horizontal="center")
            elif col_idx in (8, 9, 10):
                cell.number_format = "#,##0.00"
                cell.alignment = Alignment(horizontal="right")
            elif col_idx in (3, 4):
                cell.alignment = Alignment(horizontal="center")
    total_row = data_start + len(rows)
    total_amount = sum(r["amount"] for r in rows)
    total_vat = sum(r["vat"] for r in rows)
    total_grand = sum(r["total"] for r in rows)
    ws.merge_cells(start_row=total_row, start_column=1, end_row=total_row, end_column=2)
    ws[f"A{total_row}"] = "รวมทั้งหมด"
    ws.merge_cells(start_row=total_row, start_column=3, end_row=total_row, end_column=7)
    ws[f"C{total_row}"] = thai_baht_text(total_grand)
    ws[f"H{total_row}"] = f"=SUM(H{data_start}:H{total_row - 1})"
    ws[f"I{total_row}"] = f"=SUM(I{data_start}:I{total_row - 1})"
    ws[f"J{total_row}"] = f"=SUM(J{data_start}:J{total_row - 1})"
    for col_idx in range(1, 11):
        cell = ws.cell(row=total_row, column=col_idx)
        cell.border = border
        if col_idx in (8, 9, 10):
            cell.number_format = "#,##0.00"
            cell.font = Font(name="Angsana New", size=BODY_FONT_SIZE, bold=True)
    footer_row = total_row + 3
    ws.merge_cells(start_row=footer_row, start_column=1, end_row=footer_row, end_column=4)
    ws.cell(row=footer_row, column=1, value="ผู้วางบิล ……………………………………..")
    ws.merge_cells(start_row=footer_row, start_column=7, end_row=footer_row, end_column=9)
    ws.cell(row=footer_row, column=7, value="ผู้รับวางบิล  .................................................")
    ws.merge_cells(start_row=footer_row + 1, start_column=7, end_row=footer_row + 1, end_column=9)
    ws.cell(row=footer_row + 1, column=7, value="วันที่รับวางบิล………………………..")
    ws.cell(row=footer_row, column=10, value="วันที่วางบิล.................................")
    ws.cell(row=footer_row + 1, column=10, value="วันที่จ่ายชำระ.............................")


def write_narrow_sheet(ws, config: dict, rows: list[dict], combined_small: bool) -> None:
    apply_base_format(ws, max_rows=400)
    write_common_header(ws, config["customer_line"], config["address_line"], config["tax_line"], config["bill_no"], config.get("title"))
    border = make_border()
    fill = PatternFill("solid", fgColor="D9EAD3")
    headers = ["ลำดับ", "วันที่", "เลขที่บิล", "รหัสสาขา", "ประเภท", "จำนวน", "ราคา", "จำนวนเงิน", "ภาษีมูลค่าเพิ่ม", "เงินรวมทั้งสิ้น"]
    row_idx = 10
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row_idx, column=col_idx, value=header)
        cell.font = Font(name="Angsana New", size=BODY_FONT_SIZE, bold=True)
        cell.alignment = Alignment(horizontal="center")
        cell.border = border
        cell.fill = fill
    data_start = 11
    for row_idx, record in enumerate(rows, start=data_start):
        label = record["branch_name"] if combined_small or config["kind"] == "lawson_small_single" else "หลอดใหญ่"
        values = [
            record["seq"],
            record["date"],
            record["bill_no"],
            record["branch"],
            label,
            record["qty"],
            record["price"],
            f"=ROUND(F{row_idx}*G{row_idx}/1.07,2)",
            f"=ROUND(J{row_idx}-H{row_idx},2)",
            f"=ROUND(F{row_idx}*G{row_idx},2)",
        ]
        for col_idx, value in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = border
            if col_idx == 2:
                cell.number_format = "dd/mm/yyyy"
                cell.alignment = Alignment(horizontal="center")
            elif col_idx in (1, 6):
                cell.number_format = "0"
                cell.alignment = Alignment(horizontal="center")
            elif col_idx == 7:
                cell.number_format = "0.00"
                cell.alignment = Alignment(horizontal="center")
            elif col_idx in (8, 9, 10):
                cell.number_format = "#,##0.00"
                cell.alignment = Alignment(horizontal="right")
            elif col_idx in (3, 4):
                cell.alignment = Alignment(horizontal="center")
    total_row = data_start + len(rows)
    total_amount = sum(r["amount"] for r in rows)
    total_vat = sum(r["vat"] for r in rows)
    total_grand = sum(r["total"] for r in rows)
    ws[f"B{total_row}"] = "รวม"
    ws[f"C{total_row}"] = thai_baht_text(total_grand)
    ws[f"F{total_row}"] = f"=SUM(F{data_start}:F{total_row - 1})"
    ws[f"H{total_row}"] = f"=SUM(H{data_start}:H{total_row - 1})"
    ws[f"I{total_row}"] = f"=SUM(I{data_start}:I{total_row - 1})"
    ws[f"J{total_row}"] = f"=SUM(J{data_start}:J{total_row - 1})"
    for col_idx in range(1, 11):
        cell = ws.cell(row=total_row, column=col_idx)
        cell.border = border
        if col_idx in (8, 9, 10):
            cell.number_format = "#,##0.00"
    footer_row = total_row + 4
    ws.cell(row=footer_row, column=1, value="ลงชื่อ")
    ws.cell(row=footer_row, column=2, value="............................................... ผู้วางบิล")
    ws.cell(row=footer_row, column=7, value="ลงชื่อ  ............................................................")
    ws.cell(row=footer_row, column=10, value="ผู้รับวางบิล")
    ws.cell(row=footer_row + 2, column=1, value="วันที่วางบิล.................................")
    ws.cell(row=footer_row + 2, column=7, value="วันที่รับวางบิล .................................................")
    ws.cell(row=footer_row + 4, column=1, value="วันที่จ่ายชำระ.............................")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="สร้างไฟล์วางบิล Lawson/แฟรนไชส์ จากโฟลเดอร์ c-d หรือ SI-xxxx")
    parser.add_argument(
        "input_path",
        nargs="?",
        default=str(DEFAULT_BASE_DIR.parent),
        help="path ของโฟลเดอร์ c-d หรือโฟลเดอร์ SI-xxxx",
    )
    parser.add_argument(
        "--month",
        default="2026-04",
        help="เดือนข้อมูลรูปแบบ YYYY-MM เช่น 2026-04",
    )
    parser.add_argument(
        "--bill-date",
        default=None,
        help="วันที่วางบิลรูปแบบ YYYY-MM-DD ถ้าไม่ใส่จะใช้วันที่วันนี้",
    )
    parser.add_argument(
        "--output",
        default=None,
        help="path ไฟล์ xlsx ปลายทาง ถ้าไม่ใส่จะบันทึกใน outputs ของโฟลเดอร์ input",
    )
    parser.add_argument(
        "--sheet-group",
        choices=["all", "lawson-small", "lawson-big", "franchise-big"],
        default="all",
        help="เลือกกลุ่มชีตที่ต้องการสร้าง",
    )
    parser.add_argument(
        "--big-source",
        default="SI-*",
        help="ชื่อโฟลเดอร์หรือ path สำหรับข้อมูลหลอดใหญ่/แฟรนไชส์",
    )
    parser.add_argument(
        "--small-source",
        default="LAWSON",
        help="ชื่อโฟลเดอร์หรือ path สำหรับข้อมูลหลอดเล็ก",
    )
    return parser.parse_args()


def main() -> None:
    global BASE_DIR, OUTPUT_XLSX, BILL_DATE, PERIOD_START, PERIOD_END

    args = parse_args()
    root_input = Path(args.input_path).expanduser().resolve()
    PERIOD_START, PERIOD_END = month_bounds(args.month)

    if args.bill_date:
        bill_day = datetime.strptime(args.bill_date, "%Y-%m-%d").date()
    else:
        bill_day = date.today()
    BILL_DATE = format_thai_bill_date(bill_day)

    big_source_input = Path(args.big_source).expanduser()
    if big_source_input.is_absolute():
        big_source_dir = resolve_source_dir(big_source_input)
    else:
        big_source_dir = resolve_preferred_source_dir(root_input, args.big_source)

    small_source_input = Path(args.small_source).expanduser()
    if small_source_input.is_absolute():
        small_source_dir = resolve_source_dir(small_source_input)
    else:
        small_source_dir = resolve_preferred_source_dir(root_input, args.small_source)

    if args.output:
        OUTPUT_XLSX = Path(args.output).expanduser().resolve()
    else:
        output_root = root_input if root_input.is_dir() else SCRIPT_DIR
        output_dir = output_root / "outputs"
        filename = f"วางบิล_{args.month}_อัปเดตจาก_{big_source_dir.name}.xlsx"
        OUTPUT_XLSX = output_dir / filename

    BASE_DIR = big_source_dir
    big_customers, big_trans_rows, big_bill_rows = load_context(big_source_dir)
    small_customers, small_trans_rows, small_bill_rows = load_context(small_source_dir)
    wb = Workbook()
    default_ws = wb.active
    wb.remove(default_ws)

    if args.sheet_group == "lawson-small":
        selected_configs = [config for config in SHEET_CONFIGS if config["kind"].startswith("lawson_small")]
    elif args.sheet_group == "lawson-big":
        selected_configs = [config for config in SHEET_CONFIGS if config["kind"] == "lawson_big"]
    elif args.sheet_group == "franchise-big":
        selected_configs = [config for config in SHEET_CONFIGS if config["kind"] == "franchise_big"]
    else:
        selected_configs = SHEET_CONFIGS

    for config in selected_configs:
        ws = wb.create_sheet(config["name"])
        if config["kind"].startswith("lawson_small"):
            rows = build_small_bill_rows(small_customers, small_bill_rows, config["branch_ids"])
        else:
            rows = build_trans_rows(big_customers, big_trans_rows, config["branch_ids"], config["product_code"])
        if config["kind"] == "lawson_big":
            write_lawson_big_sheet(ws, config, rows)
        elif config["kind"] == "franchise_big":
            write_narrow_sheet(ws, config, rows, combined_small=False)
        elif config["kind"] == "lawson_small_combined":
            write_narrow_sheet(ws, config, rows, combined_small=True)
        else:
            write_narrow_sheet(ws, config, rows, combined_small=True)
        print(config["name"], len(rows), round(sum(r["total"] for r in rows), 2))

    OUTPUT_XLSX.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT_XLSX)
    print(f"big_source={big_source_dir}")
    print(f"small_source={small_source_dir}")
    print(OUTPUT_XLSX)


if __name__ == "__main__":
    main()
