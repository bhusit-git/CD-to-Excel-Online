from __future__ import annotations

import argparse
import re
import struct
from datetime import date
from datetime import datetime
from datetime import timedelta
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


SCRIPT_DIR = Path(__file__).resolve().parent
DEFAULT_BASE_DIR = SCRIPT_DIR / "LAWSON"
DEFAULT_OUTPUT_XLSX = SCRIPT_DIR / "outputs" / "lawson_หลอดใหญ่แพ็ค_เรียงวันที่เลขสาขา.xlsx"
BASE_DIR = DEFAULT_BASE_DIR
OUTPUT_XLSX = DEFAULT_OUTPUT_XLSX
PERIOD_START = "20260401"
PERIOD_END = "20260430"
THAI_MONTHS = [
    "",
    "มกราคม",
    "กุมภาพันธ์",
    "มีนาคม",
    "เมษายน",
    "พฤษภาคม",
    "มิถุนายน",
    "กรกฎาคม",
    "สิงหาคม",
    "กันยายน",
    "ตุลาคม",
    "พฤศจิกายน",
    "ธันวาคม",
]


def parse_bill_number(raw: str) -> str:
    raw = raw.strip().replace(" ", "")
    if not raw:
        return ""
    try:
        return str(int(round(float(raw))))
    except ValueError:
        return raw


def parse_ymd(raw: str) -> datetime:
    return datetime.strptime(raw.strip(), "%Y%m%d")


def format_thai_bill_date(value: date) -> str:
    return f"วันที่ {value.day} {THAI_MONTHS[value.month]} {value.year + 543}"


def month_bounds(month_text: str) -> tuple[str, str]:
    start_day = datetime.strptime(month_text + "-01", "%Y-%m-%d").date()
    next_month = (start_day.replace(day=28) + timedelta(days=4)).replace(day=1)
    end_day = next_month - timedelta(days=1)
    return start_day.strftime("%Y%m%d"), end_day.strftime("%Y%m%d")


def resolve_source_dir(input_path: Path) -> Path:
    input_path = input_path.expanduser().resolve()
    if (input_path / "MCUST.DBF").exists() and (input_path / "ATRANS.DBF").exists():
        return input_path

    candidates = []
    for child in input_path.iterdir():
        if child.is_dir() and (child / "MCUST.DBF").exists() and (child / "ATRANS.DBF").exists():
            candidates.append(child)
    if not candidates:
        raise FileNotFoundError(f"ไม่พบโฟลเดอร์ข้อมูลที่มีไฟล์ DBF ใน {input_path}")
    preferred_names = {"LAWSON", "SI-2568"}
    candidates.sort(
        key=lambda p: (p.name.upper() not in preferred_names, -p.stat().st_mtime.timestamp(), p.name.upper())
    )
    return candidates[0]


def load_dbf(path: Path) -> list[dict]:
    with path.open("rb") as f:
        hdr = f.read(32)
        numrec = struct.unpack("<I", hdr[4:8])[0]
        hlen = struct.unpack("<H", hdr[8:10])[0]
        rlen = struct.unpack("<H", hdr[10:12])[0]
        fields = []
        while True:
            desc = f.read(32)
            if not desc or desc[0] == 0x0D:
                break
            name = desc[:11].split(b"\x00", 1)[0].decode("ascii", "ignore").strip()
            typ = chr(desc[11])
            flen = desc[16]
            fields.append((name, typ, flen))
        offsets = []
        pos = 1
        for _, _, flen in fields:
            offsets.append(pos)
            pos += flen
        rows = []
        f.seek(hlen)
        for _ in range(numrec):
            rec = f.read(rlen)
            if not rec or rec[:1] == b"*":
                continue
            row = {}
            for (name, typ, flen), off in zip(fields, offsets):
                raw = rec[off : off + flen]
                if typ == "D":
                    value = raw.decode("ascii", "ignore").strip()
                else:
                    value = raw.decode("cp874", "ignore").strip(" \x00")
                row[name] = value
            rows.append(row)
    return rows


def clean_branch_name(raw: str) -> str:
    raw = raw.strip()
    raw = re.sub(r"^P\d+\s*สาขา\s*", "", raw)
    return raw


def load_rows() -> list[dict]:
    customer_rows = load_dbf(BASE_DIR / "MCUST.DBF")
    lawson_customers = {
        row["ID"]: row
        for row in customer_rows
        if "สหลอว์สัน" in row.get("NAME", "")
    }
    trans_rows = load_dbf(BASE_DIR / "ATRANS.DBF")

    rows = []
    for row in trans_rows:
        cust_id = row.get("CUST_VEND", "")
        if cust_id not in lawson_customers:
            continue
        if row.get("PROD_CODE") != "04":
            continue
        date_raw = row.get("DATE", "")
        if not (PERIOD_START <= date_raw <= PERIOD_END):
            continue

        cust = lawson_customers[cust_id]
        total = float(row.get("AMT") or 0)
        amount = round(total / 1.07, 12)
        vat = round(total - amount, 12)
        branch_name = clean_branch_name(cust.get("SH_NAME") or cust.get("SHIP_NAME") or cust.get("NAME") or "")
        rows.append(
            {
                "date": parse_ymd(date_raw),
                "branch": f"P{cust_id}",
                "bill_no": parse_bill_number(row.get("BILL_NO", "")),
                "branch_name": branch_name,
                "qty": float(row.get("QTY") or 0),
                "price": float(row.get("PRICE") or 0),
                "amount": amount,
                "vat": vat,
                "total": total,
            }
        )

    rows.sort(key=lambda r: (r["branch_name"], r["date"], r["branch"], r["bill_no"]))
    for idx, row in enumerate(rows, start=1):
        row["seq"] = idx
    return rows


THAI_DIGITS = ["ศูนย์", "หนึ่ง", "สอง", "สาม", "สี่", "ห้า", "หก", "เจ็ด", "แปด", "เก้า"]
THAI_POSITIONS = ["", "สิบ", "ร้อย", "พัน", "หมื่น", "แสน"]


def thai_baht_text(amount: float) -> str:
    integer = int(round(amount))

    def read_under_million(num: int) -> str:
        if num == 0:
            return ""
        parts = []
        digits = list(map(int, str(num)))
        size = len(digits)
        for idx, digit in enumerate(digits):
            pos = size - idx - 1
            if digit == 0:
                continue
            if pos == 0:
                if digit == 1 and size > 1:
                    parts.append("เอ็ด")
                else:
                    parts.append(THAI_DIGITS[digit])
            elif pos == 1:
                if digit == 1:
                    parts.append("สิบ")
                elif digit == 2:
                    parts.append("ยี่สิบ")
                else:
                    parts.append(THAI_DIGITS[digit] + "สิบ")
            else:
                parts.append(THAI_DIGITS[digit] + THAI_POSITIONS[pos])
        return "".join(parts)

    if integer == 0:
        return "ศูนย์บาทถ้วน"

    chunks = []
    while integer > 0:
        chunks.append(integer % 1_000_000)
        integer //= 1_000_000

    words = []
    for idx in range(len(chunks) - 1, -1, -1):
        chunk = chunks[idx]
        if chunk == 0:
            continue
        chunk_text = read_under_million(chunk)
        if idx > 0:
            words.append(chunk_text + "ล้าน")
        else:
            words.append(chunk_text)
    return "".join(words) + "บาทถ้วน"


def apply_base_format(ws):
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A12"

    col_widths = {
        "A": 7,
        "B": 12,
        "C": 14,
        "D": 12,
        "E": 14,
        "F": 30,
        "G": 24,
        "H": 10,
        "I": 11,
        "J": 12,
        "K": 13,
        "L": 14,
    }
    for col, width in col_widths.items():
        ws.column_dimensions[col].width = width

    for row in range(1, 500):
        for col in range(1, 13):
            ws.cell(row=row, column=col).font = Font(name="Angsana New", size=16)
            ws.cell(row=row, column=col).alignment = Alignment(vertical="center")


def write_header(ws, bill_date_text: str):
    ws.merge_cells("A1:L1")
    ws["A1"] = "ใบวางบิล/ใบแจ้งหนี้"
    ws["A1"].font = Font(name="Angsana New", size=20, bold=True)
    ws["A1"].alignment = Alignment(horizontal="center")

    ws.merge_cells("A2:J2")
    ws["A2"] = "บริษัท ซูเปอร์ ไอซ์ จำกัด  สำนักงานใหญ่"
    ws["A2"].font = Font(name="Angsana New", size=18, bold=True)

    ws["K2"] = "ต้นฉบับ"
    ws["K2"].font = Font(name="Angsana New", size=16, bold=True)

    ws.merge_cells("A3:H3")
    ws["A3"] = "ที่อยู่ 18/39 ซอยนวมินทร์ 111 แยก 15 แขวงนวมินทร์ เขตบึงกุ่ม กรุงเทพมหานคร 10240"
    ws.merge_cells("I3:L3")
    ws["I3"] = "เลขที่บิล 6905001"

    ws.merge_cells("A4:H4")
    ws["A4"] = "เลขประจำตัวผู้เสียภาษี 0105542031756"

    ws.merge_cells("I5:L5")
    ws["I5"] = bill_date_text
    ws["I5"].alignment = Alignment(horizontal="center")
    ws["I5"].font = Font(name="Angsana New", size=16, bold=True)

    ws.merge_cells("A6:L6")
    ws["A6"] = "ชื่อลูกค้า : บริษัท สห ลอว์สัน จำกัด"
    ws["A6"].font = Font(name="Angsana New", size=16, bold=True)

    ws.merge_cells("A7:L7")
    ws["A7"] = "เลขที่ 2170 อาคารกรุงเทพทาวเวอร์ ชั้น 3 ถนนเพชรบุรีตัดใหม่ แขวงบางกะปิ เขตห้วยขวาง กรุงเทพมหานคร 10310"

    ws.merge_cells("A8:L8")
    ws["A8"] = "เลขประจำตัวผู้เสียภาษี 105-55516-6337"

    ws.merge_cells("A9:L9")
    ws["A9"] = "รายการ : น้ำแข็งหลอดใหญ่แพ็ค 1.4 KG (ห่อ)"
    ws["A9"].font = Font(name="Angsana New", size=16, bold=True)


def write_table(ws, rows: list[dict]) -> int:
    headers = [
        "ลำดับ",
        "วันที่",
        "เลขที่บิล",
        "เลขที่สาขา",
        "เลขที่บิล",
        "สาขา",
        "ชื่อสินค้า",
        "จำนวน",
        "ราคา/หน่วย",
        "จำนวนเงิน",
        "ภาษีมูลค่าเพิ่ม",
        "เงินรวมทั้งสิ้น",
    ]

    header_fill = PatternFill("solid", fgColor="D9EAD3")
    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    header_row = 11
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=header)
        cell.font = Font(name="Angsana New", size=16, bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.fill = header_fill
        cell.border = border

    data_start = 12
    for row_idx, record in enumerate(rows, start=data_start):
        values = [
            record["seq"],
            record["date"],
            "",
            record["branch"],
            record["bill_no"],
            record["branch_name"],
            "",
            record["qty"],
            record["price"],
            record["amount"],
            record["vat"],
            record["total"],
        ]
        for col_idx, value in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = border
            if col_idx == 2:
                cell.number_format = "dd/mm/yyyy"
                cell.alignment = Alignment(horizontal="center")
            elif col_idx in (1, 8):
                cell.number_format = "0"
                cell.alignment = Alignment(horizontal="center")
            elif col_idx == 9:
                cell.number_format = "0.00"
                cell.alignment = Alignment(horizontal="center")
            elif col_idx in (10, 11, 12):
                cell.number_format = '#,##0.00'
                cell.alignment = Alignment(horizontal="right")
            elif col_idx in (3, 4, 5):
                cell.alignment = Alignment(horizontal="center")

    total_amount = sum(r["amount"] for r in rows)
    total_vat = sum(r["vat"] for r in rows)
    total_grand = sum(r["total"] for r in rows)

    total_row = data_start + len(rows)
    ws["A" + str(total_row)] = "รวมทั้งหมด"
    ws["A" + str(total_row)].font = Font(name="Angsana New", size=16, bold=True)
    ws.merge_cells(start_row=total_row, start_column=1, end_row=total_row, end_column=2)
    ws["C" + str(total_row)] = "IN64002102"
    ws["D" + str(total_row)] = thai_baht_text(total_grand)
    ws.merge_cells(start_row=total_row, start_column=4, end_row=total_row, end_column=9)
    ws["J" + str(total_row)] = total_amount
    ws["K" + str(total_row)] = total_vat
    ws["L" + str(total_row)] = total_grand

    for col_idx in range(1, 13):
        ws.cell(row=total_row, column=col_idx).border = border
        if col_idx in (10, 11, 12):
            ws.cell(row=total_row, column=col_idx).number_format = '#,##0.00'
            ws.cell(row=total_row, column=col_idx).font = Font(name="Angsana New", size=16, bold=True)

    return total_row


def write_footer(ws, total_row: int):
    footer_row = total_row + 3
    ws.merge_cells(start_row=footer_row, start_column=1, end_row=footer_row, end_column=4)
    ws.cell(row=footer_row, column=1, value="ผู้วางบิล ……………………………………..")
    ws.merge_cells(start_row=footer_row, start_column=8, end_row=footer_row, end_column=11)
    ws.cell(row=footer_row, column=8, value="ผู้รับวางบิล .................................................")

    ws.merge_cells(start_row=footer_row + 1, start_column=8, end_row=footer_row + 1, end_column=11)
    ws.cell(row=footer_row + 1, column=8, value="วันที่รับวางบิล………………………..")

    ws.merge_cells(start_row=footer_row, start_column=12, end_row=footer_row, end_column=12)
    ws.cell(row=footer_row, column=12, value="วันที่วางบิล.................................")
    ws.cell(row=footer_row + 1, column=12, value="วันที่จ่ายชำระ.............................")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="สร้างไฟล์วางบิล Lawson หลอดใหญ่")
    parser.add_argument(
        "input_path",
        nargs="?",
        default=str(DEFAULT_BASE_DIR),
        help="path ของโฟลเดอร์ LAWSON, SI-xxxx หรือโฟลเดอร์แม่",
    )
    parser.add_argument(
        "--month",
        default="2026-04",
        help="เดือนข้อมูลรูปแบบ YYYY-MM เช่น 2026-04",
    )
    parser.add_argument(
        "--bill-date",
        default=None,
        help="วันที่วางบิลรูปแบบ YYYY-MM-DD ถ้าไม่ใส่จะใช้วันที่วันนี้",
    )
    parser.add_argument(
        "--output",
        default=None,
        help="path ไฟล์ xlsx ปลายทาง",
    )
    return parser.parse_args()


def main():
    global BASE_DIR, OUTPUT_XLSX, PERIOD_START, PERIOD_END

    args = parse_args()
    root_input = Path(args.input_path).expanduser().resolve()
    BASE_DIR = resolve_source_dir(root_input)
    PERIOD_START, PERIOD_END = month_bounds(args.month)

    if args.bill_date:
        bill_day = datetime.strptime(args.bill_date, "%Y-%m-%d").date()
    else:
        bill_day = date.today()
    bill_date_text = format_thai_bill_date(bill_day)

    if args.output:
        OUTPUT_XLSX = Path(args.output).expanduser().resolve()
    else:
        output_root = root_input if root_input != BASE_DIR else BASE_DIR.parent
        OUTPUT_XLSX = output_root / "outputs" / f"lawson_หลอดใหญ่แพ็ค_{args.month}_{BASE_DIR.name}.xlsx"

    rows = load_rows()
    OUTPUT_XLSX.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "ลอสันหลอดใหญ่"

    apply_base_format(ws)
    write_header(ws, bill_date_text)
    total_row = write_table(ws, rows)
    write_footer(ws, total_row)

    wb.save(OUTPUT_XLSX)
    print(OUTPUT_XLSX)
    print(f"rows={len(rows)}")


if __name__ == "__main__":
    main()
